'''
@Project ：jingmai_uitest/data
@Author ：Aby
@Date ：2024/5/29 19:23 
@Describe:...
'''

from openpyxl import load_workbook
from pprint import pprint
from utils.log_helper.logger import logger


def filter_empty(old_l):
    """
    过滤序列中的空值
    """
    new_l = []
    for i in old_l:
        if i:
            new_l.append(i)
    return new_l


def data_by_csv():
    pass


def data_by_mysql():
    pass


# openpyxl 2.0.5以下版本才适用
def data_by_excel(file):
    """ 从excel中加载测试用例的信息 """
    wb = load_workbook(file)
    suite_dict = {}  # 以套件名称为Key，以用例为value
    quit_flag = False
    for sheet_name in wb.get_sheet_names():
        if quit_flag:
            break
        case_dict = {}  # 以名称为Key，以步骤为value的字典
        case_name = ""
        ws = wb.get_sheet_by_name(sheet_name)
        for line in ws.rows:
            line = tuple(map(lambda cell: cell.value, line))
            _id = line[0]
            if isinstance(_id, int):
                if _id == -999:
                    quit_flag = True
                    break
                if _id == -1:
                    case_name = line[3]
                    case_dict[case_name] = []  # 以用例名称为Key，创建新的空用例
                elif _id > 0:
                    case_dict[case_name].append(filter_empty(line))  # 为用例填充步骤

        for k, v in list(case_dict.items()):
            if not len(v):
                case_dict.pop(k)  # 去除空用例

        ws_info = f'工作表【{ws.title}】：包含{len(case_dict)}个用例'
        logger.info(ws_info)
        suite_dict[ws.title] = case_dict  # 本测试套件的所有用例
    return suite_dict


# ===================================================================================================================
# openpyxl最新版本适用
def data_by_excel_bk(file):
    '''
    从excel中加载测试用例的信息
    '''
    wb = load_workbook(file)
    # workbook_info = f'文件{file.stem},包含了{len(wb.worksheets)} 工作表'
    # print(workbook_info)
    # logger.debug(workbook_info)
    suite_dict = {}  # 以套件名称为Key，以用例为value
    quit_flag = False
    for ws in wb.worksheets:
        if quit_flag:
            break
        case_dict = {}  # 以名称为Key，以步骤为value的字典
        case_name = ""
        for line in ws.iter_rows(values_only=True):
            _id = line[0]
            if isinstance(_id, int):
                if _id == -999:
                    quit_flag = True
                    break
                if _id == -1:
                    case_name = line[3]
                    case_dict[case_name] = []  # 以用例名称为Key，创建新的空用例
                elif _id > 0:
                    case_dict[case_name].append(filter_empty(line))  # 为用例填充步骤

        for k, v in list(case_dict.items()):
            if not len(v):
                case_dict.pop(k)  # 去除空用例

        ws_info = f'工作表【{ws.title}】：包含{len(case_dict)}个用例'
        logger.info(ws_info)
        suite_dict[ws.title] = case_dict  # 本测试套件的所有用例
    return suite_dict


if __name__ == '__main__':
    f_name = r'D:\hjt_workspace\auto\jingmai_uitest\testcases\test_AA.xlsx'
    pprint(data_by_excel(f_name))
